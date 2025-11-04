# 研究板块运动
import datetime
from typing import List#  板块热力excel
import openpyxl
from openpyxl.styles import PatternFill

import akshare as ak
import baostock as bs

stock_code = '002342'
stock_codes = ['002342', '300961', '300810', '301063', '002278']
start_date = '20250313'
end_date = '202504010'


def get_kdata_mapping(stock_codes: List[str], start_date: str, end_date: str):
    df_mapping = {}
    for code in stock_codes:
        stock_daily_df = ak.stock_zh_a_hist(symbol=code, start_date=start_date, end_date=end_date, adjust="qfq")
        df_mapping[code] = stock_daily_df
print(bs.__version__)
# kdata_mapping = get_kdata_mapping()

# 写入xlsx文件：
# A列是日期，从start_date到end_date
# 接着从B列开始，每行两列为一个组，分别写入股票代码和涨幅




# # 加载Excel文件
# wb = openpyxl.load_workbook("data.xlsx")
# ws = wb.active
#
# # 定义颜色填充规则
# red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色
# yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
#
# # 遍历B列数据（假设数值在B列）
# for row in range(2, ws.max_row + 1):
#     cell_value = ws[f"B{row}"].value
#     if cell_value > 100:  # 大于100填充红色
#         ws[f"B{row}"].fill = red_fill
#     elif cell_value > 50:  # 50~100填充黄色
#         ws[f"B{row}"].fill = yellow_fill
#
# wb.save("colored_data.xlsx")